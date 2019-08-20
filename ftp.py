import requests, threading, random, time
from python_anticaptcha import AnticaptchaClient, NoCaptchaTaskProxylessTask
from datetime import datetime
from threading import Thread
from bs4 import BeautifulSoup
from discord_hooks import Webhook
import openpyxl
from itertools import cycle


#PRODUCT ID  
PID = '148370'

#problems with 3ds 



###global vars 
in_queue_list = []
past_queue_list = []
stock_id_list = []

##################
# Give the location of the file 
path = 'slot_kishans.xlsx'
#path = 'ftp_slot.xlsx'
# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 

sheet_obj = wb_obj.active 

max_col = sheet_obj.max_column
max_row = sheet_obj.max_row

# Will print a particular row value 
info = []
for j in range(2, max_row + 1):
    list_row = []
    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row = j, column = i) 
        list_row.append(str(cell_obj.value))
    info.append(list_row)

print(datetime.now().strftime('%T') + " Loaded:", len(info), "slots")
print(datetime.now().strftime('%T') + " PID:", PID)
#print(info)


def set_prox_for_session():

    mythread = threading.currentThread().getName()
    i = int(mythread[7:])

    try:
        PROXYFILENAME="proxy.txt"
        PROXS=open(PROXYFILENAME,'r').read().split('\n')
        PROXYLIST=[p.strip() for p in PROXS if p.strip()!='']
        PROXYLIST.append(None)
    except:
        PROXYFILENAME=None
        PROXYLIST=[None]

    try:
        prox=PROXYLIST[i-1]
    except: 
        prox=None

    #print(PROXYLIST)
    try:
        print(datetime.now().strftime('%T') + " [" + mythread + "] " + "using proxy", prox)

        prox_details = prox.split(':')

        http_prox = "http://" + prox_details[2] + ':' + prox_details[3] + '@' + prox_details[0] + ':' + prox_details[1]
        https_prox = "http://" + prox_details[2] + ':' + prox_details[3] + '@' + prox_details[0] + ':' + prox_details[1]
    
    except:
        http_prox = None
        https_prox = None


    #proxiesss = {"https": prox}
    proxiesss = {"http": http_prox, "https": https_prox}
    req = requests.Session()
    
    #added this
    #req.trust_env=False
    
    req.proxies.update(proxiesss)
    in_queue_list.append(req)

    time.sleep(1)

    Load_queue()

def Load_queue():
    mythread = threading.currentThread().getName()
    i = int(mythread[7:])

    while (in_queue_list != []):

        req = in_queue_list[i-1]

        #headers = {
            #'Upgrade-Insecure-Requests': '1',
           #'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36',
        #}

        headers = {
            'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'en-US,en;q=0.9',
        }


        #load_fp = req.get('https://www.footpatrol.com', headers=headers)
        queue_cookies = req.cookies.get_dict()
        #print(queue_cookies)

        while ('akavpau_VP1' not in queue_cookies):
        #while ("Footpatrol - Waiting Page" in load_fp.text):

            try:
                print(datetime.now().strftime('%T') + " [" + mythread + "] " + "IN queue")
                sleep = random.randint(6, 15)
                time.sleep(sleep)
                load_fp = req.get('https://www.footpatrol.com', headers=headers, timeout=35)
                #print(load_fp.text)
                load_fp.raise_for_status()
                queue_cookies = req.cookies.get_dict()
                #print(queue_cookies)

            except requests.exceptions.HTTPError as errh:
                print (datetime.now().strftime('%T') + " [" + mythread + "] " + "Http Error:",errh)
                time.sleep(10)
                req.cookies.clear()
            except requests.exceptions.ConnectionError as errc:
                print (datetime.now().strftime('%T') + " [" + mythread + "] " + "Error Connecting:",errc)
                time.sleep(10)
                req.cookies.clear()
            except requests.exceptions.Timeout as errt:
                print (datetime.now().strftime('%T') + " [" + mythread + "] " + "Timeout Error:",errt)
                time.sleep(10)
                req.cookies.clear()
            except requests.exceptions.RequestException as err:
                print (datetime.now().strftime('%T') + " [" + mythread + "] " + "OOps: Something Else",err)
                time.sleep(10)
                req.cookies.clear()
            #in_queue_list.append(req)

            #if ("Footpatrol - Waiting Page" not in load_fp.text):
                #past_queue_list.append(req)
                #in_queue_list.pop(req)

                #else:
                    #past_queue_list.append(req)
                    #pass
        #print(load_fp.text)
        #print(queue_cookies)
        #past_que_cookie = queue_cookies.get('akavpau_VP1')
        #print('akavpau_vp1:', past_que_cookie)

        #get_stock = req.get('https://www.footpatrol.com/product/yellow-adidas-originals-yeezy-350-glow/339645_footpatrolcom/stock/', headers=headers)
        #prinr(get_stock.text)

        print(datetime.now().strftime('%T') + " [" + mythread + "] " + "Past queue using proxy")
        past_queue_list.append(req)

        browse_rand_prods(req)
        
        #carting(req)
        #akavpau_vp1#1558771150~id=31bbbebb29daf5fe505e1b11e92b3a47

def browse_rand_prods(sess):

    mythread = threading.currentThread().getName()
    prod_list = ['https://www.footpatrol.com/product/black-footpatrol-london-to-paris-mug/094959_footpatrolcom/', 'https://www.footpatrol.com/product/multi-jason-markk-travel-kit/296681_footpatrolcom/', 'https://www.footpatrol.com/product/footpatrol-bar-logo-socks/073321_footpatrolcom/', 'https://www.footpatrol.com/product/multi-retaw-x-footpatrol-nimbus-sneaker-liquid-fragrance/267742_footpatrolcom/', 'https://www.footpatrol.com/product/white-jason-markk-quick-wipes-3-pack/215279_footpatrolcom/', 'https://www.footpatrol.com/product/white-footpatrol-bar-logo-socks/073323_footpatrolcom/', 'https://www.footpatrol.com/product/black-retaw-x-footpatrol-nimbus-room-tag/267741_footpatrolcom/', 'https://www.footpatrol.com/product/multi-vault-by-vans-x-ralph-steadman-skate-deck/135492_footpatrolcom/', 'https://www.footpatrol.com/product/yellow-adidas-originals-trefoil-cap/102908_footpatrolcom/', 'https://www.footpatrol.com/product/black-new-era-x-starcow-59fifty-cap/332724_footpatrolcom/', 'https://www.footpatrol.com/product/green-nikelab-acg-nrg-beanie/126641_footpatrolcom/', 'https://www.footpatrol.com/product/multi-adidas-x-bape-football/153812_footpatrolcom/', 'https://www.footpatrol.com/product/white-jason-markk-rtu-foam/085500_footpatrolcom/', 'https://www.footpatrol.com/product/jason-markk-holiday-box-gift-set/053907_footpatrolcom/', 'https://www.footpatrol.com/product/multi-jason-markk-4oz-premium-cleaning-kit/002525_footpatrolcom/', 'https://www.footpatrol.com/product/footpatrol-duffy-gasmask-ring/094870_footpatrolcom/', 'https://www.footpatrol.com/product/white-vault-by-vans-x-ralph-steadman-saw-fish-skate-deck/135486_footpatrolcom/', 'https://www.footpatrol.com/product/white-footpatrol-gas-mask-logo-socks/254205_footpatrolcom/', 'https://www.footpatrol.com/product/multi-vault-by-vans-x-ralph-steadman-black-rhino-deck/135487_footpatrolcom/', 'https://www.footpatrol.com/product/retaw-x-footpatrol-gift-set/055328_footpatrolcom/', 'https://www.footpatrol.com/product/white-jason-markk-quick-wipes-30-pack/215280_footpatrolcom/', 'https://www.footpatrol.com/product/white-jason-markk-8oz-premium-shoe-cleaner/102083_footpatrolcom/', 'https://www.footpatrol.com/product/multi-vault-by-vans-x-ralph-steadman-skate-deck/135491_footpatrolcom/', 'https://www.footpatrol.com/product/medicom-x-atmos-animal-print-bearbrick-100--400/085795_footpatrolcom/', 'https://www.footpatrol.com/product/black-footpatrol-gas-mask-logo-socks/254200_footpatrolcom/', 'https://www.footpatrol.com/product/pintrill-for-footpatrol-bar-logo-pin-badge/012027_footpatrolcom/', 'https://www.footpatrol.com/product/black-footpatrol-x-theobalds-exploration-club-souvenir-cap/097859_footpatrolcom/', 'https://www.footpatrol.com/product/retaw-x-footpatrol-nimbus-candle/267743_footpatrolcom/', 'https://www.footpatrol.com/product/white-nike-air-max-aw84-cap/103373_footpatrolcom/', 'https://www.footpatrol.com/product/grey-nike-mars-cap/180655_footpatrolcom/', 'https://www.footpatrol.com/product/medicom-x-atmos-animal-bearbrick-1000/085797_footpatrolcom/', 'https://www.footpatrol.com/product/multi-medicom-x-yasuto-sasada-peacock-berbrick/079550_footpatrolcom/', 'https://www.footpatrol.com/product/multi-medicom-homer-simpson-bebrick-1000/058449_footpatrolcom/', 'https://www.footpatrol.com/product/medicom-x-atmos-elephant-bearbrick-100--400/085789_footpatrolcom/', 'https://www.footpatrol.com/product/brown-jason-markk-premium-brush/102086_footpatrolcom/', 'https://www.footpatrol.com/product/medicom-x-have-a-good-time-berbrick-100--400/079397_footpatrolcom/', 'https://www.footpatrol.com/product/brown-jason-markk-standard-shoe-cleaning-brush/114988_footpatrolcom/', 'https://www.footpatrol.com/product/multi-jason-markk-suede-cleaning-kit/016659_footpatrolcom/', 'https://www.footpatrol.com/product/green-footpatrol-x-theobalds-exploration-club-logo-cap/097858_footpatrolcom/']

    #12 am release
    req = sess
    time_curr = datetime.now().strftime('%T')

    while ('22:00:01' < time_curr < '23:59:59'):
    #while ('22:00:01' > time_curr > '23:59:59'):
        browse = random.choice(prod_list)
        print(datetime.now().strftime('%T') + " [" + mythread + "] " + "Browsing", browse)
        atc_req = req.get(browse)
        time.sleep(15)

    print(datetime.now().strftime('%T') + " [" + mythread + "] " + "Past 12 am release - goin getting size id")

    if (stock_id_list == []):
        get_stock_id()

    else:
        print(datetime.now().strftime('%T') + " [" + mythread + "] " + "time to cook start solving capthchas")
        carting(sess)

def get_stock_id():
    if (past_queue_list != []):
        mythread = threading.currentThread().getName()
        req = past_queue_list[0]
        main_prod = 'https://www.footpatrol.com/search/' + PID + '_footpatrolcom/'

        headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'en-US,en;q=0.9',
            }

        while (stock_id_list == []):
            get_url = req.get(main_prod, headers=headers, timeout=15)
            id_url = get_url.url + 'stock/'
            get_id = req.get(id_url, headers=headers, timeout=15)

            parse_page = get_id.text

            queue_cookies = req.cookies.get_dict()
            #print(queue_cookies)

            if ('akavpau_VP1' not in queue_cookies):
                print(datetime.now().strftime('%T') + " [" + mythread + "] " + "Stock page on queue: thrown back in Queue")
                past_queue_list.pop(0)
                Load_queue()

            else:
                soup = BeautifulSoup(parse_page, 'lxml')
                get_id_buttons = soup.findAll('button',{'class':'btn btn-default '})

                for data_sku in get_id_buttons:
                    stock_id_list.append(data_sku['data-sku'])

        print(datetime.now().strftime('%T') + " [" + mythread + "] " + "stock ids loaded")

    else:
        print(datetime.now().strftime('%T') + " no sessions past queue ")


def get_captcha():
    mythread = threading.currentThread().getName()
    print(datetime.now().strftime('%T') + " [" + mythread + "] " + "Getting captcha...")

    rcap = requests.Session()

    try:
        captcha_req = rcap.get('http://cartchefs.footpatrol.com:5000/token')
        solved_captcha = captcha_req.text

        if (solved_captcha == 'ERROR'):
            ########edit  this#########
            api_key = '2b73a91bb7885de80b19051763e62e6c'
            ############################

            site_key = '6LdhYxYUAAAAAAcorjMQeKmZb6W48bqb0ZEDRPCl' # undftd 
            #site_key = '6LetKEIUAAAAAPk-uUXqq9E82MG3e40OMt_74gjS'
            #url = 'https://losangeles.doverstreetmarket.com/new-items/raffle'
            url = 'https://app.viralsweep.com'

            client = AnticaptchaClient(api_key)
            task = NoCaptchaTaskProxylessTask(url, site_key)
            job = client.createTask(task)
            job.join()
            solved_captcha = (job.get_solution_response())
            print(datetime.now().strftime('%T') + " Got captcha response")
            
    except:
        print(datetime.now().strftime('%T') + " [" + mythread + "] " + "run/solve manual harvester")
    ########edit  this#########
        api_key = '2b73a91bb7885de80b19051763e62e6c'
    ############################

        site_key = '6LfEwHQUAAAAACTyJsAICi2Lz2oweGni8QOhV-Yl' # grab from site
        url = 'https://www.footpatrol.com/'

        client = AnticaptchaClient(api_key)
        task = NoCaptchaTaskProxylessTask(url, site_key)
        job = client.createTask(task)
        job.join()
        solved_captcha = (job.get_solution_response())
        print(datetime.now().strftime('%T') + " [" + mythread + "] " + "Got captcha response")
        
    return solved_captcha


def carting(past_que):
    mythread = threading.currentThread().getName()
    i = int(mythread[7:])


    try:
        k = i-1
        slot_info = info[k]
    except: 
        k = i % len(info)
        slot_info = info[k]

    main = True

    while (main == True):
        success = False


        print(datetime.now().strftime('%T') + " [" + mythread + "] " + "carting")

        firstName = slot_info[0]
        lastName = slot_info[1]
        phone = slot_info[2]
        address1 = slot_info[3]
        address2 = slot_info[4]
        if address2 == 'None':
            address2 = ""
        city = slot_info[5]
        state = slot_info[6]
        postcode = slot_info[7]
        cardNumber = slot_info[8]
        cardHolderName = slot_info[9]
        expiryMonth = slot_info[10]
        if len(expiryMonth) == 1:
            expiryMonth = '0' + expiryMonth
        expiryYear = slot_info[11]
        cvcCode = slot_info[12]
        email = slot_info[13]

        req = past_que
        #req = 
        captcha = get_captcha()

        sess_cookies = req.cookies.get_dict()
        #print(sess_cookies)

        abck_value = sess_cookies.get('_abck', '')

        headers = {
            'Referer': 'https://www.footpatrol.com/',
            'Origin': 'https://www.footpatrol.com',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
            'Content-Type': 'text/plain;charset=UTF-8',
        }

        data = '{"sensor_data":"7a74G7m23Vrp0o5c9080891.41-1,2,-94,-100,Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36,uaend,12147,20030107,en-US,Gecko,3,0,0,0,384313,8656999,1680,1010,1680,1050,641,902,1565,,cpen:0,i1:0,dm:0,cwen:0,non:1,opc:0,fc:0,sc:0,wrc:1,isc:0,vib:1,bat:1,x11:0,x12:1,8319,0.984134980492,780974328499,loc:-1,2,-94,-101,do_en,dm_en,t_en-1,2,-94,-105,0,0,0,0,960,113,0;0,-1,0,0,1589,488,0;-1,2,-94,-102,0,0,0,0,960,113,0;0,-1,0,0,1589,488,0;-1,2,-94,-108,-1,2,-94,-110,-1,2,-94,-117,-1,2,-94,-111,-1,2,-94,-109,-1,2,-94,-114,-1,2,-94,-103,-1,2,-94,-112,https://www.footpatrol.com/-1,2,-94,-115,1,1,0,0,0,0,0,3,0,1561948656998,-999999,16709,0,0,2784,0,0,7,0,0,' + abck_value + ',27656,-1,-1,30261693-1,2,-94,-106,0,0-1,2,-94,-119,-1-1,2,-94,-122,0,0,0,0,1,0,0-1,2,-94,-123,-1,2,-94,-70,-1-1,2,-94,-80,94-1,2,-94,-116,8657028-1,2,-94,-118,73921-1,2,-94,-121,;6;-1;0"}'

        #print(data)
        response = req.post('https://www.footpatrol.com/resources/c0b6940c232082f9f5742adced4807', headers=headers, data=data)

        sess_cookies = req.cookies.get_dict()
        abck_value = sess_cookies.get('_abck', '')

        if '~0~' in abck_value:
            print(datetime.now().strftime('%T') + " [" + mythread + "] " + "got abck validated carting now...")

        #print(sess_cookies)

            headers = {

                'origin': 'https://www.footpatrol.com',
                'accept-encoding': 'gzip, deflate, br',
                'accept-language': 'en-US,en;q=0.9',
                'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
                'content-type': 'application/json',
                'accept': '*/*',
                'x-requested-with': 'XMLHttpRequest',
        
                #'origin': 'https://www.footpatrol.com',
                #'accept-encoding': 'gzip, deflate, br',
                #'accept-language': 'en-US,en;q=0.9',
                #'x-requested-with': 'XMLHttpRequest',
                #'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36',
                #'content-type': 'application/json',
                #'accept': '*/*',
            }

            data = '{"customisations":false,"cartPosition":null,"recaptchaResponse":"'+ captcha +'","cartProductNotification":null,"quantityToAdd":1}'

            #sizes = ['339645_footpatrolcom.001366582', '339645_footpatrolcom.001366583', '339645_footpatrolcom.001366584', '339645_footpatrolcom.001366586','339645_footpatrolcom.001366587', '339645_footpatrolcom.001366588', '339645_footpatrolcom.001366589','339645_footpatrolcom.001366588', '339645_footpatrolcom.001366590', '339645_footpatrolcom.001366591', '339645_footpatrolcom.001366592', '339645_footpatrolcom.001366593', '339645_footpatrolcom.001366594', '339645_footpatrolcom.001366595', '339645_footpatrolcom.001366596', '339645_footpatrolcom.001366597']
            carting_size = random.choice(stock_id_list)
            cart_url = 'https://www.footpatrol.com/cart/' + carting_size

            atc_req = req.post(cart_url, headers=headers, data=data)

            if (atc_req.status_code == 200):
                print(datetime.now().strftime('%T') + " [" + mythread + "] " + "ATC successfull. guest checkout with email")

                data = '{"email":"' + email + '"}'
                email_post = req.post('https://www.footpatrol.com/checkout/guest/', data=data)

                if (email_post.status_code == 200):
                    print(datetime.now().strftime('%T') + " [" + mythread + "] " + "submitting address")

                    #data = '{"useDeliveryAsBilling":true,"country":"Netherlands|nl","locale":"","firstName":"kaay","lastName":"meer","phone":"5611145001","address1":"41 Vrouwenzandstraa pkg 1","address2":"","town":"Purmerend","county":"","postcode":"1443WG","addressPredict":"","setOnCart":"deliveryAddressID"}'
                    data = '{"useDeliveryAsBilling":true,"country":"United States|us","locale":"","firstName":"' + firstName + '","lastName":"' + lastName + '","phone":"' + phone + '","address1":"' + address1 + '","address2":"' + address2 +'","town":"' + city + '","county":"' + state + '","postcode":"' + postcode + '","addressPredict":"","setOnCart":"deliveryAddressID"}'

                    addy_req = req.post('https://www.footpatrol.com/myaccount/addressbook/add/', data=data)

                    #print(addy_req.text)

                    if (addy_req.status_code == 201):

                        payment_status = False
                        while(payment_status == False):

                            print(datetime.now().strftime('%T') + " [" + mythread + "] " + "picking card")

                            data = {
                              'paySelect': 'card'
                            }
                            select_card_req = req.post('https://www.footpatrol.com/checkout/paymentV3/', data=data)

                            if (select_card_req.status_code == 200):

                                html_doc = select_card_req.text
                                soup = BeautifulSoup(html_doc, 'lxml')
                                ayden_link = soup.find('iframe').get('src')
                                print(datetime.now().strftime('%T') + " [" + mythread + "] " + "checking out")
                                
                                #print(ayden_link)
                                #req2 = requests.Session()
                                get_payment_page = req.get(ayden_link)
                                #cookies = requests.utils.dict_from_cookiejar(req2.cookies)
                                #print(cookies)

                                parse_page = get_payment_page.text
                                soup = BeautifulSoup(parse_page, 'lxml')

                                sig = soup.find('input', {'name':'sig'})['value']
                                merchantReference = soup.find('input', {'name':'merchantReference'})['value']
                                #brandCode = soup.find('input', {'name':'brandCode'})['value']
                                paymentAmount = soup.find('input', {'name':'paymentAmount'})['value']
                                currencyCode = soup.find('input', {'name':'currencyCode'})['value']
                                shipBeforeDate = soup.find('input', {'name':'shipBeforeDate'})['value']
                                skinCode = soup.find('input', {'name':'skinCode'})['value']
                                merchantAccount = soup.find('input', {'name':'merchantAccount'})['value']
                                shopperLocale = soup.find('input', {'name':'shopperLocale'})['value']
                                stage = soup.find('input', {'name':'stage'})['value']
                                sessionId = soup.find('input', {'name':'sessionId'})['value']
                                sessionValidity = soup.find('input', {'name':'sessionValidity'})['value']
                                shopperEmail = soup.find('input', {'name':'shopperEmail'})['value']
                                shopperReference = soup.find('input', {'name':'shopperReference'})['value']
                                recurringContract = soup.find('input', {'name':'recurringContract'})['value']
                                resURL = soup.find('input', {'name':'resURL'})['value']
                                allowedMethods = soup.find('input', {'name':'allowedMethods'})['value']
                                blockedMethods = soup.find('input', {'name':'blockedMethods'})['value']
                                originalSession = (soup.find('input', {'name':'originalSession'})['value']) + ':bGl2ZS5hZHllbi5jb20='
                                billingAddress_street = soup.find('input', {'name':'billingAddress.street'})['value']
                                billingAddress_houseNumberOrName = soup.find('input', {'name':'billingAddress.houseNumberOrName'})['value']
                                billingAddress_city = soup.find('input', {'name':'billingAddress.city'})['value']
                                billingAddress_postalCode = soup.find('input', {'name':'billingAddress.postalCode'})['value']
                                billingAddress_stateOrProvince = soup.find('input', {'name':'billingAddress.stateOrProvince'})['value']
                                billingAddress_country = soup.find('input', {'name':'billingAddress.country'})['value']
                                billingAddressType = soup.find('input', {'name':'billingAddressType'})['value']
                                billingAddressSig = soup.find('input', {'name':'billingAddressSig'})['value']
                                deliveryAddress_street = soup.find('input', {'name':'deliveryAddress.street'})['value']
                                deliveryAddress_houseNumberOrName = soup.find('input', {'name':'deliveryAddress.houseNumberOrName'})['value']
                                deliveryAddress_city = soup.find('input', {'name':'deliveryAddress.city'})['value']
                                deliveryAddress_postalCode = soup.find('input', {'name':'deliveryAddress.postalCode'})['value']
                                deliveryAddress_stateOrProvince = soup.find('input', {'name':'deliveryAddress.stateOrProvince'})['value']
                                deliveryAddress_country = soup.find('input', {'name':'deliveryAddress.country'})['value']
                                deliveryAddressType = soup.find('input', {'name':'deliveryAddressType'})['value']
                                deliveryAddressSig = soup.find('input', {'name':'deliveryAddressSig'})['value']
                                shopper_firstName = soup.find('input', {'name':'shopper.firstName'})['value']
                                shopper_lastName = soup.find('input', {'name':'shopper.lastName'})['value']
                                merchantIntegration_type = soup.find('input', {'name':'merchantIntegration.type'})['value']
                                dfValue = soup.find('input', {'name':'dfValue'})['value']
                                usingFrame = soup.find('input', {'name':'usingFrame'})['value']
                                usingPopUp = soup.find('input', {'name':'usingPopUp'})['value']
                                shopperBehaviorLog = '{"numberBind":"1","holderNameBind":"1","cvcBind":"1","deactivate":"6","activate":"6","numberFieldFocusCount":"2","numberFieldLog":"fo@700,cl@701,bl@726,fo@762,KU@772,KU@779,KL@782,ch@788,bl@788","numberFieldClickCount":"1","numberFieldBlurCount":"2","numberFieldKeyCount":"3","numberUnkKeysFieldLog":"91@772,91@779","numberFieldChangeCount":"1","numberFieldEvHa":"total=0","holderNameFieldFocusCount":"1","holderNameFieldLog":"fo@788,cl@789,KL@802,KL@804,KL@805,KL@807,KL@808,KL@809,Ks@812,KL@814,KL@815,KL@817,KL@819,KL@821,KL@822,KL@823,KL@825,ch@832,bl@832","holderNameFieldClickCount":"1","holderNameFieldKeyCount":"15","holderNameFieldChangeCount":"1","holderNameFieldEvHa":"total=0","holderNameFieldBlurCount":"1","cvcFieldFocusCount":"2","cvcFieldLog":"fo@873,cl@873,bl@894,fo@916,KN@930,KN@935,KN@938","cvcFieldClickCount":"1","cvcFieldBlurCount":"1","cvcFieldKeyCount":"3"}'


                                headers = {
                                    'Connection': 'keep-alive',
                                    'Cache-Control': 'max-age=0',
                                    'Origin': 'https://live.adyen.com',
                                    'Upgrade-Insecure-Requests': '1',
                                    'Content-Type': 'application/x-www-form-urlencoded',
                                    'x-requested-with': 'XMLHttpRequest',
                                    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36',
                                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
                                    'Referer': ayden_link,
                                    'Accept-Encoding': 'gzip, deflate, br',
                                    'Accept-Language': 'en-US,en;q=0.9',
                                }

                                data = {    
                                  'displayGroup': 'card',
                                  'card.cardNumber': cardNumber,
                                  'card.cardHolderName': cardHolderName,
                                  'card.expiryMonth': expiryMonth,
                                  'card.expiryYear': expiryYear,
                                  'card.cvcCode': cvcCode,
                                  'sig': sig,
                                  'merchantReference': merchantReference,
                                  'brandCode':  'brandCodeUndef',
                                  'paymentAmount': paymentAmount,
                                  'currencyCode': currencyCode,
                                  'shipBeforeDate': shipBeforeDate,
                                  'skinCode': skinCode,
                                  'merchantAccount': merchantAccount,
                                  'shopperLocale': shopperLocale,
                                  'stage': stage,
                                  'sessionId': sessionId,
                                  'sessionValidity': sessionValidity,
                                  'shopperEmail': shopperEmail,
                                  'shopperReference': shopperReference,
                                  'recurringContract': recurringContract,
                                  'resURL': 'https://www.footpatrol.com/checkout/landing/',
                                  'allowedMethods': allowedMethods,
                                  'blockedMethods': blockedMethods,
                                  'originalSession': originalSession,
                                  'billingAddress.street': billingAddress_street,
                                  'billingAddress.houseNumberOrName': billingAddress_houseNumberOrName,
                                  'billingAddress.city': billingAddress_city,
                                  'billingAddress.postalCode': billingAddress_postalCode,
                                  'billingAddress.stateOrProvince': billingAddress_stateOrProvince,
                                  'billingAddress.country': billingAddress_country,
                                  'billingAddressType': billingAddressType,
                                  'billingAddressSig': billingAddressSig,
                                  'deliveryAddress.street': deliveryAddress_street,
                                  'deliveryAddress.houseNumberOrName': deliveryAddress_houseNumberOrName,
                                  'deliveryAddress.city': deliveryAddress_city,
                                  'deliveryAddress.postalCode': deliveryAddress_postalCode,
                                  'deliveryAddress.stateOrProvince': deliveryAddress_stateOrProvince,
                                  'deliveryAddress.country': deliveryAddress_country,
                                  'deliveryAddressType': deliveryAddressType,
                                  'deliveryAddressSig': deliveryAddressSig,
                                  'shopper.firstName': shopper_firstName,
                                  'shopper.lastName': shopper_lastName,
                                  'merchantIntegration.type': merchantIntegration_type,
                                  'referrerURL': 'https://www.footpatrol.com/checkout/billing/',
                                  'usingFrame': 'true',
                                  'usingPopUp': usingPopUp,
                                  'shopperBehaviorLog': '{"numberBind":"1","holderNameBind":"1","cvcBind":"1","activate":"2","numberFieldFocusCount":"1","numberFieldLog":"fo@61,cl@62,KU@64,KL@65,ch@71,bl@71","numberFieldClickCount":"1","numberFieldKeyCount":"2","numberUnkKeysFieldLog":"91@64","numberFieldChangeCount":"1","numberFieldEvHa":"total=0","numberFieldBlurCount":"1","holderNameFieldFocusCount":"2","holderNameFieldLog":"fo@71,cl@72,KL@85,KL@86,KL@88,Ks@90,KL@92,KL@93,KL@94,KL@96,ch@111,bl@111,fo@145,bl@151","holderNameFieldClickCount":"1","holderNameFieldKeyCount":"8","holderNameFieldChangeCount":"1","holderNameFieldEvHa":"total=0","holderNameFieldBlurCount":"2","deactivate":"1","cvcFieldFocusCount":"1","cvcFieldLog":"fo@181,cl@182,KN@195,KN@195,KN@201,ch@255,bl@255","cvcFieldClickCount":"1","cvcFieldKeyCount":"3","cvcFieldChangeCount":"1","cvcFieldEvHa":"total=0","cvcFieldBlurCount":"1"}'
                                  }

                                cc_submit = req.post('https://live.adyen.com/hpp/completeCard.shtml', headers=headers, data=data)
                                #print("")
                                #time.sleep(2)
                                #print(cc_submit.url)

                                if ("https://www.footpatrol.com/checkout/landing/?authResult=REFUSED" in cc_submit.url):
                                    print(datetime.now().strftime('%T') + " [" + mythread + "] " + "Could not go to 3Ds")
                                    payment_status = False

                                else:
                                    parse_page = cc_submit.text
                                    soup = BeautifulSoup(parse_page, 'lxml')

                                    pageform = soup.find('form', {'id':'pageform'})['action']
                                    PaReq = soup.find('input', {'name':'PaReq'})['value']
                                    TermUrl = soup.find('input', {'name':'TermUrl'})['value']
                                    MD = soup.find('input', {'name':'MD'})['value']


                                    data = {
                                      'PaReq': PaReq,
                                      'TermUrl': 'https://live.adyen.com/hpp/complete3dIntermediate.shtml',
                                      'MD': MD,
                                      'shopperBehaviorLog': ''
                                    }

                                    submit_3d = req.post(pageform, data=data)

                                    time.sleep(5)
                                    #print(submit_3d.text)

                                    if ("https://aacsw.3ds.verifiedbyvisa.com/aacs/pahandler?" in submit_3d.url):
                                        parse_page = submit_3d.text
                                        soup = BeautifulSoup(parse_page, 'lxml')

                                        PaRes = soup.find('input', {'name':'PaRes'})['value']
                                        MD = soup.find('input', {'name':'MD'})['value']

                                        data = {
                                          'PaRes': PaRes,
                                          'MD': MD
                                        }

                                        submit_3d2 = req.post('https://live.adyen.com/hpp/complete3dIntermediate.shtml', data=data)
                                        #time.sleep(10)


                                        if ("https://live.adyen.com/hpp/complete3dIntermediate.shtml" in submit_3d2.url):
                                            parse_page = submit_3d2.text
                                            soup = BeautifulSoup(parse_page, 'lxml')

                                            MD = soup.find('input', {'name':'MD'})['value']
                                            PaRes = soup.find('input', {'name':'PaRes'})['value']

                                            data = {
                                              'MD': MD,
                                              'PaRes': PaRes
                                            }

                                            submit_3d3 = req.post('https://live.adyen.com/hpp/complete3d.shtml', data=data)
                                            #time.sleep(10)

                                            #print(submit_3d3.text)
                                            #print(submit_3d3.url)
                                            #print(submit_3d3.status_code)

                                            if ("https://www.footpatrol.com/checkout/landing/?authResult=REFUSED" in submit_3d3.url):
                                                print(datetime.now().strftime('%T') + " [" + mythread + "] " + "Card declined or something failed with 3ds")
                                                payment_status = True

                                            else:
                                                print(datetime.now().strftime('%T') + " [" + mythread + "] " + "Cook check email")
                                                payment_status = True
                                                #print(submit_3d3.text)
                                                #print(submit_3d3.url)
                                                #print(submit_3d3.status_code)
                                        else:
                                            print(datetime.now().strftime('%T') + " [" + mythread + "] " + "Error in 3ds")
                                            print(submit_3d2.url)

                                    else:
                                        print(datetime.now().strftime('%T') + " [" + mythread + "] " + "Error in 3ds")
                                        print(submit_3d.url)


                                    #success = True
                                    #main = False
                                    #print(cc_submit.url)
                                    #print(datetime.now().strftime('%T') + " [" + mythread + "] " + "checkedout check email") 

                            else:
                                print(select_card_req.text)
                                print(datetime.now().strftime('%T') + " [" + mythread + "] " + "Out stock")

                    else:
                        print(addy_req.text)
                        print(datetime.now().strftime('%T') + " [" + mythread + "] " + "error submitting address")


                else:
                    print(email_post.text)
                    print(datetime.now().strftime('%T') + " [" + mythread + "] " + "error picking guest checkout with email")

            else:
                print(atc_req.text)
                print(datetime.now().strftime('%T') + " [" + mythread + "] " + "ATC error")
        else:
            print(abck_value)
            print(datetime.now().strftime('%T') + " [" + mythread + "] " + "Invalid abck // change sensor data")


threads = []

try:
    print(datetime.now().strftime('%T') + "---------------------------------------------")
    print(datetime.now().strftime('%T') + "   Developed by @iamkishann. FOOTPATROL.PY   ")
    print(datetime.now().strftime('%T') + "---------------------------------------------\n")
    tasks = input(datetime.now().strftime('%T') + " Enter how many threads would you like to runproxy:slot 1:1: ")

    for i in range(int(tasks)):
        t = threading.Thread(target=set_prox_for_session)
        threads.append(t)
        t.start()

except (KeyboardInterrupt):
        print(datetime.now().strftime('%T') + "Program Exited.")
        quit()
